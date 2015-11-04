#! /usr/bin/python
#### map2assets.py   10/21/2015
####  M. Canazon


###script to see if assets from the Qualys Map (IPs or DNS) are in CMDB, and print it out to new excel book
###compare.py will combine all output into a single sheet &
### noMatch2compare.py will generate a list of assets that are not in the CMDB

import os
import subprocess
import openpyxl
from openpyxl import load_workbook

print os.getcwd() ###the excel books must be in the same directory as te scripts.

out = load_workbook('outTest.xlsx', data_only=True)  #this is the excel book that data is returned to.
print out.get_sheet_names()

#create sheets to list comparable data
outSheet = out.create_sheet(index=0, title = 'IP2IP') #if IP is in IP of CMDB
shortSheet = out.create_sheet(index=1, title='shortSheet') #if dns is in ShortName of CMDB
aliasSheet = out.create_sheet(index=2, title='aliasSheet') #if dns is in AliasName of CMDB
descSheet = out.create_sheet(index=3, title='descSheet') #if dns is in descirption of CMDB
print out.get_sheet_names()

#create sheet column headers
outSheet['A1'] = 'IP'
outSheet['B1'] = 'DNS'
outSheet['C1'] = 'Owner'
outSheet['D1'] = 'CI Name'
outSheet['E1'] = 'CI Alias'
outSheet['F1'] = 'CI Description'
outSheet['G1'] = 'CI IP'
outSheet['H1'] = 'CI ID'

shortSheet['A1'] = 'IP'
shortSheet['B1'] = 'DNS'
shortSheet['C1'] = 'Owner'
shortSheet['D1'] = 'CI Name'
shortSheet['E1'] = 'CI Alias'
shortSheet['F1'] = 'CI Description'
shortSheet['G1'] = 'CI IP'
shortSheet['H1'] = 'CI ID'

aliasSheet['A1'] = 'IP'
aliasSheet['B1'] = 'DNS'
aliasSheet['C1'] = 'Owner'
aliasSheet['D1'] = 'CI Name'
aliasSheet['E1'] = 'CI Alias'
aliasSheet['F1'] = 'CI Description'
aliasSheet['G1'] = 'CI IP'
aliasSheet['H1'] = 'CI ID'

descSheet['A1'] = 'IP'
descSheet['B1'] = 'DNS'
descSheet['C1'] = 'Owner'
descSheet['D1'] = 'CI Name'
descSheet['E1'] = 'CI Alias'
descSheet['F1'] = 'CI Description'
descSheet['G1'] = 'CI IP'
descSheet['H1'] = 'CI ID'

#load workbooks and sheets to pull data from
maP = load_workbook('mapTest.xlsx', data_only=True) #qualys map
asset = load_workbook('assetTest.xlsx', data_only=True) #asset inventory

mapIP = maP.get_sheet_by_name('IPs') #make sure that sheets are named correctly
assetSheet = asset.get_sheet_by_name('Assets')

#variable top store the highest row number
mRow = str(mapIP.get_highest_row())
eRow = str(assetSheet.get_highest_row())


i = 2 #variable for row number output, skips the first row (b/c it is the header row, duh)

#create data by comparing IP in map to IP in CMDB
for mapIpRow in mapIP['A1':'A' + mRow]:
		for mapIpCell in mapIpRow:
			for assetIpRow in assetSheet['E1':'E' + eRow]:
				for assetIpCell in assetIpRow:
					assetIp = assetIpCell.value.split('\n') if assetIpCell.value else ''
					mapIp = str(mapIpCell.value)
					if mapIp in assetIp:
						outSheet['A' + str(i)].value = mapIp
						print(mapIp) #just for feedback that the program is running
						dnsM = mapIP['B' + str(mapIpCell.row)].value
						owner = assetSheet['F' + str(assetIpCell.row)].value
						dnsQ = assetSheet['B' + str(assetIpCell.row)].value #cishort
						dnsQ2 = assetSheet['C' + str(assetIpCell.row)].value #cialias
						dnsQ3 = assetSheet['D' + str(assetIpCell.row)].value #ciDesc
						ciIP = assetSheet['E' + str(assetIpCell.row)].value #ciIP
						ciID = assetSheet['A' + str(assetIpCell.row)].value #ciID
						outSheet['B' + str(i)].value = dnsM
						outSheet['C' + str(i)].value = owner
						outSheet['D' + str(i)].value = dnsQ
						outSheet['E' + str(i)].value = dnsQ2
						outSheet['F' + str(i)].value = dnsQ3
						outSheet['G' + str(i)].value = ciIP
						outSheet['H' + str(i)].value = ciID
						print owner #just for feedback that the program is running
						i = i + 1
					else:
						pass

print ''
i = 2 #reset variable for row number output
print i #just for feedback that the program is running

#shortName sheet to compare Qualys DNS to CMDB CI_ShortName 
for dnsRow in mapIP['B1':'B' + mRow]:
		for dnsCell in dnsRow:
			for ciNameRow in assetSheet['B1':'B' + eRow]:
				for ciNameCell in ciNameRow:
					ciName = ciNameCell.value.split() if ciNameCell.value else ''   
					dnsName = str(dnsCell.value)
					if dnsName in ciName: 
						mapIp = mapIP['A' + str(dnsCell.row)].value
						dnsM = mapIP['B' + str(dnsCell.row)].value
						owner = assetSheet['F' + str(ciNameCell.row)].value
						dnsQ = assetSheet['B' + str(ciNameCell.row)].value #cishort
						dnsQ2 = assetSheet['C' + str(ciNameCell.row)].value #cialias
						dnsQ3 = assetSheet['D' + str(ciNameCell.row)].value #ciDesc
						ciIP = assetSheet['E' + str(ciNameCell.row)].value #ciIP
						ciID = assetSheet['A' + str(ciNameCell.row)].value #ciID
						shortSheet['A' + str(i)].value = mapIp
						shortSheet['B' + str(i)].value = dnsM
						shortSheet['C' + str(i)].value = owner
						shortSheet['D' + str(i)].value = dnsQ
						shortSheet['E' + str(i)].value = dnsQ2
						shortSheet['F' + str(i)].value = dnsQ3
						shortSheet['G' + str(i)].value = ciIP
						shortSheet['H' + str(i)].value = ciID
						print dnsM #feedback
						i = i + 1
					else:
						pass

print '' 
i = 2 #reset variable for row number output
print i #feedback

#shortName sheet to compare Qualys DNS to CMDB CI_AliasName
for dnsRow in mapIP['B1':'B' + mRow]:
		for dnsCell in dnsRow:
			for ciAliasRow in assetSheet['C1':'C' + eRow]:
				for ciAliasCell in ciAliasRow:
					ciAlias = ciAliasCell.value.split() if ciAliasCell.value else ''   
					dnsName = str(dnsCell.value)
					if dnsName in ciAlias:
						mapIp = mapIP['A' + str(dnsCell.row)].value
						dnsM = mapIP['B' + str(dnsCell.row)].value
						owner = assetSheet['F' + str(ciAliasCell.row)].value
						dnsQ = assetSheet['B' + str(ciAliasCell.row)].value #cishort
						dnsQ2 = assetSheet['C' + str(ciAliasCell.row)].value #cialias
						dnsQ3 = assetSheet['D' + str(ciAliasCell.row)].value #ciDesc
						ciIP = assetSheet['E' + str(ciAliasCell.row)].value #ciIP
						ciID = assetSheet['A' + str(ciAliasCell.row)].value #ciID
						aliasSheet['A' + str(i)].value = mapIp
						aliasSheet['B' + str(i)].value = dnsM
						aliasSheet['C' + str(i)].value = owner
						aliasSheet['D' + str(i)].value = dnsQ
						aliasSheet['E' + str(i)].value = dnsQ2
						aliasSheet['F' + str(i)].value = dnsQ3
						aliasSheet['G' + str(i)].value = ciIP
						aliasSheet['H' + str(i)].value = ciID
						print dnsQ2 #feedback
						i = i + 1
					else:
						pass

#test print
print ''
i = 2
print i #feedback

##shortName sheet to compare Qualys DNS to CMDB CI_Description				
for dnsRow in mapIP['B1':'B' + mRow]:
		for dnsCell in dnsRow:
			for ciDescRow in assetSheet['D1':'D' + eRow]:
				for ciDescCell in ciDescRow:
					ciDesc = ciDescCell.value.split() if ciDescCell.value else ''   
					dnsName = str(dnsCell.value)
					if dnsName in ciDesc:
						mapIp = mapIP['A' + str(dnsCell.row)].value
						dnsM = mapIP['B' + str(dnsCell.row)].value
						owner = assetSheet['F' + str(ciDescCell.row)].value
						dnsQ = assetSheet['B' + str(ciDescCell.row)].value #cishort
						dnsQ2 = assetSheet['C' + str(ciDescCell.row)].value #cialias
						dnsQ3 = assetSheet['D' + str(ciDescCell.row)].value #ciDesc
						ciIP = assetSheet['E' + str(ciDescCell.row)].value #ciIP
						ciID = assetSheet['A' + str(ciDescCell.row)].value #ciID
						descSheet['A' + str(i)].value = mapIp
						descSheet['B' + str(i)].value = dnsM
						descSheet['C' + str(i)].value = owner
						descSheet['D' + str(i)].value = dnsQ
						descSheet['E' + str(i)].value = dnsQ2
						descSheet['F' + str(i)].value = dnsQ3
						descSheet['G' + str(i)].value = ciIP
						descSheet['H' + str(i)].value = ciID
						
						print dnsM
						print dnsQ3 #feedback
						print ' '
						i = i + 1
					else:
						pass

print 'Map2Assets DONE' #feedback
					
out.save('outTest.xlsx')

#launch next script, compare.py

subprocess.Popen(['/usr/bin/python', 'compare.py'])
out.save('outTest.xlsx')

print "Monkeys" #cause why not! 
