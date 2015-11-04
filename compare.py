#! /usr/bin/python
#### 10/21/2015	compare.py
#### M. Canazon
###for comparing excel sheets to see if the same rows exist in multiple sheets

import os
import subprocess
import openpyxl
from openpyxl import load_workbook

print os.getcwd()

#get work book
out = load_workbook('outTest.xlsx', data_only=True)
print out.get_sheet_names()

################################
###out.remove_sheet(out.get_sheet_by_name('COMPARE')) 
#this is for easily testing the code and not having to delete the sheet manually when testing
#if you want to use this, uncomment that line
################################

#create sheet in outTest book to output data
compareSheet = out.create_sheet(index=0, title = 'COMPARE')

#get sheets
outSheet = out.get_sheet_by_name('IP2IP')
shortSheet = out.get_sheet_by_name('shortSheet')
aliasSheet = out.get_sheet_by_name('aliasSheet')
descSheet = out.get_sheet_by_name('descSheet')
compareSheet = out.get_sheet_by_name('COMPARE')
print out.get_sheet_names()

#create column headers in new COMPARE sheet
compareSheet['A1'] = 'IP'
compareSheet['B1'] = 'DNS'
compareSheet['C1'] = 'Owner'
compareSheet['D1'] = 'CI Name'
compareSheet['E1'] = 'CI Alias'
compareSheet['F1'] = 'CI Description'
compareSheet['G1'] = 'CI IP'
compareSheet['H1'] = 'CI ID'


#create str variables of highest row #
oRow = outSheet.get_highest_row()
sRow = shortSheet.get_highest_row()
aRow = aliasSheet.get_highest_row()
dRow = descSheet.get_highest_row()
cRow = compareSheet.get_highest_row()


#variable for new row to be inserted into
i = cRow
i = i + 1 #sjip the first line(b/c of headers, duh!)
print i #curiousity + feedback

compdict = {} # compare[dnsM]

#print IP2IP sheet rows to COMPARE sheet
 #this does not need to be compared, because the COMPARE sheet is empty
for mapIpCell in range(2, (oRow + 1)):
	mapIp = outSheet['A' + str(mapIpCell)].value
	owner = outSheet['C' + str(mapIpCell)].value
	dnsM = outSheet['B' + str(mapIpCell)].value
	dnsQ = outSheet['D' + str(mapIpCell)].value #cishort
	dnsQ2 = outSheet['E' + str(mapIpCell)].value #cialias
	dnsQ3 = outSheet['F' + str(mapIpCell)].value #ciDesc
	ciIP = outSheet['G' + str(mapIpCell)].value #ciIP
	ciID = outSheet['H' + str(mapIpCell)].value #ciID
	
	compdict.setdefault(str(dnsM)) #add dns to dictionary
	compareSheet['A' + str(i)].value = mapIp
	compareSheet['B' + str(i)].value = dnsM
	compareSheet['C' + str(i)].value = owner
	compareSheet['D' + str(i)].value = dnsQ
	compareSheet['E' + str(i)].value = dnsQ2
	compareSheet['F' + str(i)].value = dnsQ3
	compareSheet['G' + str(i)].value = ciIP
	compareSheet['H' + str(i)].value = ciID
	print(mapIp + 'added') #feedback
	i = i + 1

print ' '	
print i	
out.save('outTest.xlsx')
print 'Comparing Short Sheet...'

#if shortsheet [A] is not in compdict sheet, print row to compare sheet 
for mIpCell in range(2, (sRow + 1)):		
	#variables to add shortsheet data to compare sheet 
	maIp = shortSheet['A' + str(mIpCell)].value
	dnsM = shortSheet['B' + str(mIpCell)].value
	owner = shortSheet['C' + str(mIpCell)].value
	dnsQ = shortSheet['D' + str(mIpCell)].value #cishort
	dnsQ2 = shortSheet['E' + str(mIpCell)].value #cialias
	dnsQ3 = shortSheet['F' + str(mIpCell)].value #ciDesc
	ciIP = shortSheet['G' + str(mIpCell)].value #ciIP
	ciID = shortSheet['H' + str(mIpCell)].value #ciID
	if str(dnsQ) in compdict:
		pass
	else:
		compdict.setdefault(str(dnsQ)) #add dns to dictionary
		compareSheet['A' + str(i)].value = maIp
		compareSheet['B' + str(i)].value = dnsM
		compareSheet['C' + str(i)].value = owner
		compareSheet['D' + str(i)].value = dnsQ
		compareSheet['E' + str(i)].value = dnsQ2
		compareSheet['F' + str(i)].value = dnsQ3
		compareSheet['G' + str(i)].value = ciIP
		compareSheet['H' + str(i)].value = ciID
		print(dnsQ + 'added')
		i = i + 1	


print 'done' 
print ''		
out.save('outTest.xlsx')

###################
#if alias [A] is not in COMPARE [A] sheet, print row to compare sheet 
for mapIpCell in range(2, (aRow + 1)):
	mapIp = aliasSheet['A' + str(mapIpCell)].value 
	dnsM = aliasSheet['B' + str(mapIpCell)].value #mapdns
	owner = aliasSheet['C' + str(mapIpCell)].value
	dnsQ = aliasSheet['D' + str(mapIpCell)].value #cishort
	dnsQ2 = aliasSheet['E' + str(mapIpCell)].value #cialias
	dnsQ3 = aliasSheet['F' + str(mapIpCell)].value #ciDesc
	ciIP = aliasSheet['G' + str(mapIpCell)].value #ciIP
	ciID = aliasSheet['H' + str(mapIpCell)].value #ciID	
	if str(dnsQ2) in compdict:		
		pass
	else:
		compdict.setdefault(str(dnsQ2)) #add dns to dictionary
		#add IP data to compare sheet 
		compareSheet['A' + str(i)].value = mapIp
		compareSheet['B' + str(i)].value = dnsM
		compareSheet['C' + str(i)].value = owner
		compareSheet['D' + str(i)].value = dnsQ
		compareSheet['E' + str(i)].value = dnsQ2
		compareSheet['F' + str(i)].value = dnsQ3
		compareSheet['G' + str(i)].value = ciIP
		compareSheet['H' + str(i)].value = ciID
		print dnsQ2 
		i = i + 1	
print''
out.save('outTest.xlsx')
			
#if descsheet [A] is not in COMPARE [A] sheet, print row to compare sheet 
for mIpCell in range(2, (dRow + 1)):
	mIp = descSheet['A' + str(mIpCell)].value
	dnsM = descSheet['B' + str(mIpCell)].value
	owner = descSheet['C' + str(mIpCell)].value
	dnsQ = descSheet['D' + str(mIpCell)].value #cishort
	dnsQ2 = descSheet['E' + str(mIpCell)].value #cialias
	dnsQ3 = descSheet['F' + str(mIpCell)].value #ciDesc
	ciIP = descSheet['G' + str(mIpCell)].value #ciIP
	ciID = descSheet['H' + str(mIpCell)].value #ciID
	
	if dnsQ3 in compdict:	
		pass
	else:	
		compdict.setdefault(str(dnsQ3)) #add dns to dictionary	
		#add IP data to compare sheet 
		compareSheet['A' + str(i)].value = mIp
		compareSheet['B' + str(i)].value = dnsM
		compareSheet['C' + str(i)].value = owner
		compareSheet['D' + str(i)].value = dnsQ
		compareSheet['E' + str(i)].value = dnsQ2
		compareSheet['F' + str(i)].value = dnsQ3
		compareSheet['G' + str(i)].value = ciIP
		compareSheet['H' + str(i)].value = ciID
		print dnsQ3
		i = i + 1	

print 'done' 
			
out.save('outTest.xlsx')


#launch next script, noMatch2compare.py
subprocess.Popen(['/usr/bin/python', 'noMatch2compare.py'])
print "TOTALLY DONE"
out.save('outTest.xlsx')


print '     ***** Unicorn Fantastic! You did it! *****'			

