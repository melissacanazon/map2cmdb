#! /usr/bin/python
#### 10/21/2015	noMatch2compare.py
#### M. Canazon
###if map IP does not match CMDB IP, this creates a list to give to Asset Management to add the assets to the CMDB 

import os
import openpyxl
from openpyxl import load_workbook
import xlrd
print os.getcwd()

#load workbooks and sheets
maP = load_workbook('mapTest.xlsx', data_only=True)
compare = load_workbook('outTest.xlsx', data_only=True)

ip2ip = compare.get_sheet_by_name('IP2IP')
mapIP = maP.get_sheet_by_name('IPs')
matchSheet = maP.get_sheet_by_name('NoIPmatch')
checkSheet = maP.get_sheet_by_name('Check')
compareSheet = compare.get_sheet_by_name('COMPARE')



print maP.get_sheet_names()
print compare.get_sheet_names()

#create sheet column headers
 #if IP is in IPs
matchSheet['A1'] = 'IP'
matchSheet['B1'] = 'DNS'
matchSheet['C1'] = 'Owner'
matchSheet['D1'] = 'CI_Name'
matchSheet['E1'] = 'CI_Alias'
matchSheet['F1'] = 'CI_Description'
matchSheet['G1'] = 'CI_IP'
matchSheet['H1'] = 'CI_ID'

checkSheet['A1'] = 'IP'
checkSheet['B1'] = 'NoDups'

#create str variables of highest row #
iRow = str(ip2ip.get_highest_row())
mRow = str(mapIP.get_highest_row())
cRow = str(compareSheet.get_highest_row())

#variable for new row to be inserted into
i = checkSheet.get_highest_row()
i = i + 1
print i

# get IP from IP2IP sheet, add it to checkSheet
for mapIpRow in ip2ip['A2':('A' + iRow)]:
	for mapIpCell in mapIpRow:
		mapIp = str(mapIpCell.value)
		checkSheet['A' + str(i)].value = mapIp
		print mapIp
		i = i + 1
			
print ''		

kRow = str(checkSheet.get_highest_row())
i = 2

#del duplicates from checksheet
for mapIpRow in checkSheet['A2':('A' + kRow)]:
	for mapIpCell in mapIpRow:
		mapIp = str(mapIpCell.value)
		duPIp = str(checkSheet['A' + str(str(mapIpCell.row + 1))].value)
		if mapIp in duPIp:
			pass
		else:
			checkSheet['B' + str(i)].value = mapIp
			i = i +1
		print mapIp

print ''		

i = 2

#if checksheet duPIp matches and IP in mapTest IP sheet, pass
	#else: store IP in noIPmatch

checkdict = {} # checkdata[IP][NoDups]
for row in range(2, checkSheet.get_highest_row() + 1):		
	chkIP = checkSheet['B' + str(row)].value

	checkdict.setdefault(str(chkIP)) 

print 'dict filled \n'

for mapIpRow in range(2, mapIP.get_highest_row()):
	mapIp = mapIP['A' + str(mapIpRow)].value
	dnsM = mapIP['B' + str(mapIpRow)].value
	if mapIp in checkdict:
		pass
	else:
		print mapIp
		matchSheet['A' + str(i)].value = mapIp
		matchSheet['B' + str(i)].value = dnsM	
		i = i + 1
		
i = 2
matRow = str(matchSheet.get_highest_row())
#if IP in noIPmatch sheet is in COMPARE sheet, add details.  else: i+=1		
for matchIpRow in matchSheet['A1':'A' + matRow]:
	for matchIP in matchIpRow:
		for compareIp in compareSheet['A1':'A' + cRow]:
			for compIP in compareIp:
				matchIp = str(matchIP.value)
				coIP = str(compIP.value)
				if matchIp in coIP:
					owner = compareSheet['C' + str(compIP.row)].value
					dnsQ = compareSheet['D' + str(compIP.row)].value
					dnsQ2 = compareSheet['E' + str(compIP.row)].value
					dnsQ3 = compareSheet['F' + str(compIP.row)].value
					ciIP = compareSheet['G' + str(compIP.row)].value
					ciID = compareSheet['H' + str(compIP.row)].value
								
					matchSheet['C' + str(matchIP.row)].value = owner
					matchSheet['D' + str(matchIP.row)].value = dnsQ
					matchSheet['E' + str(matchIP.row)].value = dnsQ2
					matchSheet['F' + str(matchIP.row)].value = dnsQ3
					matchSheet['G' + str(matchIP.row)].value = ciIP
					matchSheet['H' + str(matchIP.row)].value = ciID
					i = i + 1
	
print ''		

print 'Great Scott, you did it!'	


maP.save('noMatches.xlsx')
